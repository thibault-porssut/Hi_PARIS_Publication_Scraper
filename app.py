import streamlit as st
import os
import pandas as pd
import requests
import chromedriver_autoinstaller
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import urllib.parse
from openpyxl import load_workbook
from openpyxl.styles import Font
import io
import time

st.set_page_config(page_title="Hi! PARIS Publication Scraper", layout="wide")

st.title("Hi! PARIS Publication Scraper")

# Function to sanitize filenames
def sanitize_filename(name):
    """Clean a string to make it a valid filename."""
    return "".join(c for c in name if c.isalnum() or c in (' ', '.', '_')).rstrip()

# Sidebar for file uploads and configuration
with st.sidebar:
    st.header("Configuration")
    
    # Excel file upload
    uploaded_excel = st.file_uploader("Upload Liste-affilié Excel file", type=['xlsx'])
    
    # Conference URLs
    st.subheader("Conference URLs")
    
    # Quick access to major conferences
    st.write("### Quick Conference Access")
    conf_year = st.number_input("Conference Year", min_value=2020, max_value=2030, value=2025)
    
    def validate_conference_year(url, year):
        # Check if the URL contains a different year
        import re
        years = re.findall(r'/20\d{2}/', url)
        if years and int(years[0].strip('/')) != year:
            return False
        return True
    
    col1, col2 = st.columns(2)
    
    if col1.button("Add ICCV"):
        iccv_url = f"https://iccv.thecvf.com/virtual/{conf_year}/papers.html?layout=mini&filter=author&search="
        if validate_conference_year(iccv_url, conf_year):
            if 'conference_urls' not in st.session_state:
                st.session_state.conference_urls = []
            if iccv_url not in st.session_state.conference_urls:
                st.session_state.conference_urls.append(iccv_url)
                st.success(f"Added ICCV {conf_year}")
        else:
            st.error(f"Invalid year {conf_year} for ICCV URL")
    
    if col2.button("Add ICML"):
        icml_url = f"https://icml.cc/virtual/{conf_year}/papers.html?layout=mini&filter=author&search="
        if validate_conference_year(icml_url, conf_year):
            if 'conference_urls' not in st.session_state:
                st.session_state.conference_urls = []
            if icml_url not in st.session_state.conference_urls:
                st.session_state.conference_urls.append(icml_url)
                st.success(f"Added ICML {conf_year}")
        else:
            st.error(f"Invalid year {conf_year} for ICML URL")
    
    # Show current conference URLs
    if 'conference_urls' in st.session_state and st.session_state.conference_urls:
        st.write("### Current Conference URLs:")
        for url in st.session_state.conference_urls:
            col1, col2 = st.columns([0.9, 0.1])
            col1.write(url)
            if col2.button("🗑️", key=f"delete_{url}"):
                st.session_state.conference_urls.remove(url)
                st.rerun()
    
    # Manual URL input
    st.write("### Manual URL Input")
    url_input_method = st.radio("How do you want to add more URLs?", 
                               ["Text Input", "File Upload"])
    
    if url_input_method == "Text Input":
        new_urls = st.text_area("Enter additional URLs (one per line)",
                               "").split('\n')
        if new_urls and new_urls[0]:  # Only process if there's actual input
            if 'conference_urls' not in st.session_state:
                st.session_state.conference_urls = []
            st.session_state.conference_urls.extend([url for url in new_urls if url.strip()])
    else:
        uploaded_urls = st.file_uploader("Upload conference URLs file", type=['txt'])
        if uploaded_urls:
            if 'conference_urls' not in st.session_state:
                st.session_state.conference_urls = []
            new_urls = [line.decode('utf-8').strip() for line in uploaded_urls.readlines() if line.strip()]
            st.session_state.conference_urls.extend(new_urls)
    
    # Use the conference_urls from session state
    conference_urls = st.session_state.conference_urls if 'conference_urls' in st.session_state else []

    # Configure options
    st.subheader("Scraping Options")
    wait_time = st.slider("Wait time for page load (seconds)", 1, 10, 2)
    headless = st.checkbox("Run in headless mode", value=True)

# Initialize session state for button control
if 'scraping_state' not in st.session_state:
    st.session_state.scraping_state = 'initial'  # can be 'initial', 'running', 'paused'
if 'progress_data' not in st.session_state:
    st.session_state.progress_data = {
        'current_step': 0,
        'total_steps': 0,
        'grouped_data': [],
        'pubs_seen': set(),
        'current_conf_idx': 0,
        'current_author_idx': 0
    }

# Main content area
if uploaded_excel is not None and conference_urls:
    col1, col2 = st.columns(2)
    
    if st.session_state.scraping_state == 'initial':
        if col1.button("Start Scraping"):
            st.session_state.scraping_state = 'running'
            st.session_state.progress_data = {
                'current_step': 0,
                'total_steps': 0,
                'grouped_data': [],
                'pubs_seen': set(),
                'current_conf_idx': 0,
                'current_author_idx': 0
            }
            st.rerun()
    
    elif st.session_state.scraping_state == 'running':
        if col1.button("Stop"):
            st.session_state.scraping_state = 'paused'
            st.rerun()
    
    elif st.session_state.scraping_state == 'paused':
        if col1.button("Resume"):
            st.session_state.scraping_state = 'running'
            st.rerun()
        if col2.button("Reset"):
            st.session_state.scraping_state = 'initial'
            st.rerun()

    # Only proceed with scraping if we're in running state
    if st.session_state.scraping_state == 'running':
        # Setup progress tracking
        progress_bar = st.progress(0)
        status_text = st.empty()
        results_area = st.empty()
        
        try:
            # Read Excel file
            df_affiliates = pd.read_excel(uploaded_excel, header=0)
            # Column names are already correct from the Excel file, no need to rename
            
            # Convert numeric values to strings and clean up
            df_affiliates['Last Name'] = df_affiliates['Last Name'].astype(str).str.strip()
            df_affiliates['First Name'] = df_affiliates['First Name'].astype(str).str.strip()
            
            # Remove rows where either name is NaN or empty
            df_affiliates = df_affiliates[
                (df_affiliates['Last Name'] != 'nan') & 
                (df_affiliates['First Name'] != 'nan') &
                (df_affiliates['Last Name'] != '') & 
                (df_affiliates['First Name'] != '')
            ]
            
            # Create author names
            authors = (df_affiliates['First Name'] + ' ' + df_affiliates['Last Name']).unique().tolist()
            
            status_text.write(f"Found {len(authors)} authors to process")
            
            # Configure Selenium for cloud environment
            from selenium.webdriver.chrome.service import Service
            
            chrome_options = webdriver.ChromeOptions()
            chrome_options.add_argument("--headless")
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--disable-gpu")
            chrome_options.binary_location = "/usr/bin/chromium"
            
            try:
                # Use the system-installed chromedriver
                service = Service(executable_path="/usr/bin/chromedriver")
                driver = webdriver.Chrome(service=service, options=chrome_options)
                
            except Exception as e:
                st.error(f"Failed to initialize Chrome driver: {str(e)}")
                st.error("Please try refreshing the page or contact support if the issue persists.")
                raise e
            
            grouped_data = []
            pubs_seen = set()
            
            # Initialize or get progress tracking variables
            if st.session_state.progress_data['total_steps'] == 0:
                st.session_state.progress_data['total_steps'] = len(conference_urls) * len(authors)
                st.session_state.progress_data['grouped_data'] = grouped_data
                st.session_state.progress_data['pubs_seen'] = pubs_seen
            else:
                grouped_data = st.session_state.progress_data['grouped_data']
                pubs_seen = st.session_state.progress_data['pubs_seen']
            
            total_steps = st.session_state.progress_data['total_steps']
            current_step = st.session_state.progress_data['current_step']
            
            # Resume from where we left off
            for conf_idx in range(st.session_state.progress_data['current_conf_idx'], len(conference_urls)):
                conf_url = conference_urls[conf_idx]
                conf_name = conf_url.split('/')[2]
                
                for author_idx in range(st.session_state.progress_data['current_author_idx'], len(authors)):
                    author = authors[author_idx]
                    
                    # Update progress tracking
                    st.session_state.progress_data['current_conf_idx'] = conf_idx
                    st.session_state.progress_data['current_author_idx'] = author_idx
                    # Check if we should stop
                    if st.session_state.scraping_state != 'running':
                        # Save progress
                        st.session_state.progress_data['current_step'] = current_step
                        st.session_state.progress_data['grouped_data'] = grouped_data
                        st.session_state.progress_data['pubs_seen'] = pubs_seen
                        st.stop()
                    
                    current_step += 1
                    progress = current_step / total_steps
                    progress_bar.progress(progress)
                    
                    status_text.write(f"Processing {author} for {conf_name} ({current_step}/{total_steps})")
                    
                    # Update progress in session state
                    st.session_state.progress_data['current_step'] = current_step
                    
                    search_url = f"{conf_url}{urllib.parse.quote_plus(author)}"
                    max_retries = 1
                    retry_count = 0
                    success = False
                    
                    while retry_count < max_retries and not success:
                        try:
                            st.write(f"Attempt {retry_count + 1} for {author}")
                            driver.get(search_url)
                            time.sleep(2)  # Give the page a moment to start loading
                            
                            wait = WebDriverWait(driver, wait_time)
                            paper_elements = wait.until(
                                EC.presence_of_all_elements_located((By.CSS_SELECTOR, "h5.card-title"))
                            )
                            author_elements = driver.find_elements(By.CSS_SELECTOR, "h6.card-subtitle")
                            
                            # Process found papers
                            temp_publications = []
                            for paper_el, author_el in zip(paper_elements, author_elements):
                                title = paper_el.text.strip()
                                all_authors = author_el.text.strip()
                                key = (title, conf_name)
                                if key not in pubs_seen:
                                    pubs_seen.add(key)
                                    hi_paris_in_paper = [a for a in authors if a.lower() in all_authors.lower()]
                                    temp_publications.append({
                                        "Conference": "ICML 2025",
                                        "Title": title,
                                        "Hi! PARIS Authors": ", ".join(hi_paris_in_paper),
                                        "All Authors": all_authors
                                    })
                            
                            # Look for PDFs on arXiv if publications were found
                            if temp_publications:
                                for pub in temp_publications:
                                    try:
                                        arxiv_search_url = f"https://arxiv.org/search/?query={urllib.parse.quote_plus(pub['Title'])}&searchtype=all&source=header"
                                        driver.get(arxiv_search_url)
                                        wait = WebDriverWait(driver, wait_time)
                                        pdf_link_element = wait.until(EC.presence_of_element_located((By.LINK_TEXT, "pdf")))
                                        pdf_url = pdf_link_element.get_attribute('href')
                                    except Exception:
                                        pdf_url = "N/A"
                                        st.warning(f"Could not find PDF for: {pub['Title']}")
                                    
                                    pub["Paper"] = pdf_url
                                    grouped_data.append(pub)
                            
                            success = True  # Mark this attempt as successful
                            
                        except Exception as e:
                            retry_count += 1
                            if retry_count == max_retries:
                                st.warning(f"No publications found for {author} after {max_retries} attempts.")
                            else:
                                st.warning(f"Attempt {retry_count} failed for {author}. Retrying in {retry_count * 2} seconds...")
                                time.sleep(retry_count * 2)
                                continue
            
            driver.quit()
            
            # Create DataFrame and Excel file
            if grouped_data:
                df_final = pd.DataFrame(grouped_data)
                
                # Create Excel file in memory
                output = io.BytesIO()
                df_final.to_excel(output, index=False, engine='openpyxl')
                
                # Add formatting
                wb = load_workbook(filename=output)
                ws = wb.active
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                
                output.seek(0)
                wb.save(output)
                output.seek(0)
                
                # Display results
                st.success("Scraping completed!")
                st.write("### Results Preview")
                st.dataframe(df_final)
                
                # Download button
                st.download_button(
                    label="Download Excel file",
                    data=output.getvalue(),
                    file_name="publications_HI_PARIS.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.warning("No publications found")
                
        except Exception as e:
            st.error(f"An error occurred: {str(e)}")
            if 'driver' in locals():
                driver.quit()
else:
    st.info("Please upload an Excel file and provide conference URLs to start scraping.")
